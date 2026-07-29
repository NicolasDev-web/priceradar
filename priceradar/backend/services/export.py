import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Empreendimento

logger = logging.getLogger(__name__)

HEADER_BG = "1A1A2E"
HEADER_FG = "FFFFFF"
ROW_ALT_BG = "F9F9F9"

COR_VERDE = "2E9E5B"
COR_AMARELO = "E0B010"
COR_VERMELHO = "E63E3E"

COLUNAS = [
    ("Empreendimento", 32),
    ("Anúncio", 38),
    ("Construtora", 18),
    ("Cidade", 14),
    ("Bairro", 20),
    ("Preço (R$)", 15),
    ("Área m²", 10),
    ("Preço/m²", 12),
    ("Quartos", 9),
    ("Banheiros", 10),
    ("Vagas", 7),
    ("Portal", 12),
    ("URL", 40),
    ("Data Coleta", 20),
]


def gerar_excel(empreendimentos: list[Empreendimento], preco_m2_medio: float) -> bytes:
    """Gera planilha Excel formatada e retorna bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PriceRadar"

    # Cabeçalho
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_font = Font(color=HEADER_FG, bold=True, name="Calibri", size=11)

    for col_idx, (nome_col, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 22

    limite_verde = preco_m2_medio * 0.9
    limite_vermelho = preco_m2_medio * 1.1

    # Linhas de dados
    for row_idx, emp in enumerate(empreendimentos, start=2):
        is_alt = (row_idx % 2 == 0)
        row_fill = PatternFill(fill_type="solid", fgColor=ROW_ALT_BG) if is_alt else None

        valores = [
            emp.nome_empreendimento or emp.nome_anuncio,
            emp.nome_anuncio,
            emp.construtora or "",
            emp.cidade,
            emp.bairro or "",
            emp.preco,
            emp.area_m2,
            emp.preco_m2,
            emp.quartos if emp.quartos is not None else "",
            emp.banheiros if emp.banheiros is not None else "",
            emp.vagas if emp.vagas is not None else "",
            emp.portal,
            emp.url_anuncio,
            emp.data_coleta.strftime("%d/%m/%Y %H:%M") if emp.data_coleta else "",
        ]

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            if row_fill:
                cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")

        # Formatação condicional para Preço/m² (coluna 8)
        preco_m2_cell = ws.cell(row=row_idx, column=8)
        if emp.preco_m2 < limite_verde:
            cor = COR_VERDE
        elif emp.preco_m2 > limite_vermelho:
            cor = COR_VERMELHO
        else:
            cor = COR_AMARELO
        preco_m2_cell.fill = PatternFill(fill_type="solid", fgColor=cor)
        preco_m2_cell.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)

        # Formatar preço, área e preço/m² como números (colunas 6, 7, 8)
        ws.cell(row=row_idx, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=8).number_format = 'R$ #,##0.00'

    # Linha de médias/totais
    total_row = len(empreendimentos) + 2
    ws.cell(row=total_row, column=1, value="TOTAIS / MÉDIAS").font = Font(bold=True, name="Calibri")
    ws.cell(row=total_row, column=6, value=f"=AVERAGE(F2:F{total_row-1})").number_format = 'R$ #,##0.00'
    ws.cell(row=total_row, column=7, value=f"=AVERAGE(G2:G{total_row-1})").number_format = '#,##0.00'
    ws.cell(row=total_row, column=8, value=preco_m2_medio).number_format = 'R$ #,##0.00'

    footer_fill = PatternFill(fill_type="solid", fgColor="DDEEFF")
    for col_idx in range(1, len(COLUNAS) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = footer_fill
        cell.font = Font(bold=True, name="Calibri", size=10)

    # Freeze pane no cabeçalho
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
