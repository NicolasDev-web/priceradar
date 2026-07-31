"""Testes das colunas de localização na planilha.

Latitude, Longitude e Precisão foram postas no FIM das colunas de propósito: a
formatação condicional e os formatos numéricos endereçam preço, área e preço/m²
por índice fixo (6, 7 e 8). Quem inserir uma coluna no meio pinta a célula
errada, e o teste `test_preco_area_e_preco_m2_seguem_nas_colunas_6_7_e_8` é o
que avisa.

"Precisão" existe porque uma coordenada sozinha não diz se é o endereço do
imóvel ou o centro do bairro — e quem abre a planilha não tem como saber.
"""
import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from models import Empreendimento  # noqa: E402
from services.export import COLUNAS, gerar_excel  # noqa: E402


def empreendimento(**extra) -> Empreendimento:
    base = dict(
        id="1",
        nome_anuncio="Apartamento 2 quartos",
        nome_empreendimento="Residencial Teste",
        cidade="Fortaleza",
        bairro="Meireles",
        portal="vivareal",
        preco=400_000.0,
        area_m2=60.0,
        preco_m2=6_666.67,
        construtora=None,
        quartos=2,
        banheiros=1,
        vagas=1,
        descricao=None,
        url_anuncio="https://exemplo.com/imovel/1",
        data_coleta=datetime(2026, 7, 31, 10, 0),
    )
    base.update(extra)
    return Empreendimento(**base)


def planilha(emps, media=7_000.0):
    return openpyxl.load_workbook(io.BytesIO(gerar_excel(emps, media))).active


def cabecalho(ws) -> list:
    return [c.value for c in ws[1]]


# --------------------------------------------------------------------------
# Posição das colunas
# --------------------------------------------------------------------------

def test_cabecalho_termina_com_latitude_longitude_e_precisao():
    ws = planilha([empreendimento()])
    assert cabecalho(ws)[-3:] == ["Latitude", "Longitude", "Precisão"]


def test_preco_area_e_preco_m2_seguem_nas_colunas_6_7_e_8():
    """A invariante que a formatação condicional assume por índice fixo."""
    ws = planilha([empreendimento()])
    assert cabecalho(ws)[5:8] == ["Preço (R$)", "Área m²", "Preço/m²"]
    assert [COLUNAS[i][0] for i in (5, 6, 7)] == ["Preço (R$)", "Área m²", "Preço/m²"]


def test_preco_m2_continua_recebendo_a_cor():
    """Se as colunas deslizarem, a cor vai para a célula errada."""
    ws = planilha([empreendimento(preco_m2=1_000.0)], media=7_000.0)
    assert ws.cell(row=2, column=8).fill.fgColor.rgb.endswith("2E9E5B")  # verde


# --------------------------------------------------------------------------
# Conteúdo das colunas novas
# --------------------------------------------------------------------------

def test_coordenada_vai_para_a_planilha():
    ws = planilha([empreendimento(latitude=-3.74631, longitude=-38.479013,
                                  origem_coordenada="exata")])
    assert ws.cell(row=2, column=15).value == pytest.approx(-3.74631)
    assert ws.cell(row=2, column=16).value == pytest.approx(-38.479013)


@pytest.mark.parametrize("origem,rotulo", [
    ("exata", "Endereço"),
    ("aproximada_portal", "Aproximada (portal)"),
    ("centroide_bairro", "Centro do bairro"),
])
def test_rotulo_de_precisao_por_origem(origem, rotulo):
    ws = planilha([empreendimento(latitude=-3.7, longitude=-38.5, origem_coordenada=origem)])
    assert ws.cell(row=2, column=17).value == rotulo


def test_origem_desconhecida_vira_celula_vazia():
    """Rótulo inventado seria pior que nenhum: some sem afirmar nada."""
    ws = planilha([empreendimento(latitude=-3.7, longitude=-38.5,
                                  origem_coordenada="formato_novo_do_portal")])
    assert ws.cell(row=2, column=17).value in (None, "")


def test_sem_coordenada_deixa_a_celula_vazia_e_nao_zero():
    """0,0 é uma coordenada válida — no Golfo da Guiné. Vazio é a verdade."""
    ws = planilha([empreendimento()])
    assert ws.cell(row=2, column=15).value in (None, "")
    assert ws.cell(row=2, column=16).value in (None, "")
    assert ws.cell(row=2, column=17).value in (None, "")


def test_mistura_de_origens_na_mesma_planilha():
    emps = [
        empreendimento(id="1", preco_m2=6_000.0, latitude=-3.72, longitude=-38.50,
                       origem_coordenada="exata"),
        empreendimento(id="2", preco_m2=7_000.0, latitude=-3.85, longitude=-38.49,
                       origem_coordenada="centroide_bairro"),
        empreendimento(id="3", preco_m2=8_000.0),
    ]
    ws = planilha(emps)
    assert [ws.cell(row=r, column=17).value for r in (2, 3, 4)] == [
        "Endereço", "Centro do bairro", None,
    ]


def test_linha_de_totais_cobre_as_colunas_novas():
    ws = planilha([empreendimento(latitude=-3.7, longitude=-38.5, origem_coordenada="exata")])
    # 1 empreendimento -> totais na linha 3; a faixa colorida vai até a última coluna
    assert ws.cell(row=3, column=len(COLUNAS)).fill.fgColor.rgb.endswith("DDEEFF")
